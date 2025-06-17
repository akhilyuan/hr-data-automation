import os
import pandas as pd
from config.config import FileConfig, EmployeeType
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from typing import Optional, Dict
from dateutil.relativedelta import relativedelta

class ReportGenerator:
    def __init__(self, file_config: FileConfig):
        self.file_config = file_config
        self.current_month = (datetime.now() - relativedelta(months=1)).strftime('%Y%m')
        os.makedirs(file_config.output_folder, exist_ok=True)

    def load_merge_data(self, data_path: str) -> Optional[pd.DataFrame]:
        """加载合并文件"""
        try:
            if not os.path.exists(data_path):
                print('文件不存在')
                return None
            
            df = pd.read_excel(data_path)
            print(f'成功加载数据{len(df)}行')
            return df
        
        except Exception as e:
            print(f'加载数据时出错: {e}')
            return None
        
    def generate_summary_stats(self, df: pd.DataFrame) -> dict:
        """生成数据统计"""
        stats = {
            '总人数': len(df),
            '性别结构': df['性别'].value_counts().to_dict() if '性别' in df.columns else {},
            '学历结构': df['学历分组'].value_counts().to_dict() if '学历分组' in df.columns else {},
            '年龄结构': df['年龄段'].value_counts().to_dict() if '年龄段' in df.columns else {},
            '用工性质': df['用工性质'].value_counts().to_dict() if '用工性质' in df.columns else {},
        }

        # 合同制员工统计
        if '用工性质' in df.columns:
            contract_data = df[df['用工性质'] == EmployeeType.CONTRACT.value]
            stats['合同制员工'] = {
                '总人数': len(contract_data),
                '性别结构': contract_data['性别'].value_counts().to_dict() if len(contract_data) > 0 else {},
                '学历结构': contract_data['学历分组'].value_counts().to_dict() if len(contract_data) > 0 else {},
                '年龄结构': contract_data['年龄段'].value_counts().to_dict() if len(contract_data) > 0 else {},
            }
        
        return stats
    
    def create_excel_report(self, stats: Dict) -> str:
        wb = Workbook()
        wb.remove(wb.active) # TODO 为什么要删除默认工作表


        try:
            self._create_summary_sheet(wb, stats)
            self._create_department_sheet(wb, stats)

            output_file = os.path.join(self.file_config.output_folder, f'用工月报_{self.current_month}.xlsx')
            wb.save(output_file)
            print(f'报告已保存到: {output_file}')
            # TODO 为什么要返回文件路径
            return output_file
        except Exception as e:  
            print(f'DEBUG 错误信息: {e}')
            # TODO 啥意思
            import traceback
            traceback.print_exc()
            raise

    
    # TODO 在方法面前加下划线是什么意思
    def _create_summary_sheet(self, workbook: Workbook, stats: Dict):
        ws = workbook.create_sheet('用工总体情况')

        print(f"DEBUG: current_month for title: {self.current_month}, type: {type(self.current_month)}")

        # 检查数据类型
        try:
            title = f'{self.current_month[:4]}年{self.current_month[4:]}月总体用工情况'
            print(f"DEBUG: title created successfully: {title}")
            ws['A1'] = title
        except Exception as e:
            print(f"DEBUG: Error creating title: {e}")
            raise


        # 设置标题
        ws['A1'] = f'{self.current_month[:4]}年{self.current_month[4:]}月总体用工情况'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        row = 3

        structure_items = [
            ('性别结构', stats['性别结构']),
            ('学历结构', stats['学历结构']),
            ('年龄结构', stats['年龄结构']),
            ('用工性质', stats['用工性质']),
            ('部门结构', stats.get('部门结构', {})),
            ('一线人员', stats.get('一线人员', {}))
        ]

        for index, (title, data) in enumerate(structure_items):
            if not data:
                continue
            ws[f'A{row}'] = f'{index + 1}、{title}'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                if stats['总人数'] > 0:
                    ws[f'C{row}'] = f'{(value/stats["总人数"]):.1%}'
                row += 1
            row += 1

        contract_stats = stats.get('合同制', {})
        if contract_stats.get('总人数', 0) > 0:
            ws[f'A{row}'] = '二、合同制员工结构分析'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            ws[f'A{row}'] = f'合同制员工总人数: {contract_stats["总人数"]}人'
            ws[f'A{row}'].font = Font(bold=True)
            row += 2

            constract_structure_items = [
                ('性别结构', contract_stats['性别结构']),
                ('学历结构', contract_stats['学历结构']),
                ('年龄结构', contract_stats['年龄结构']),
            ]
            for index, (title, data) in enumerate(constract_structure_items):
                ws[f'A{row}'] = f'{index + 1}、{title}'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                for key, value in data.items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = value
                    if contract_stats['总人数'] > 0:
                        ws[f'C{row}'] = f'{(value/contract_stats["总人数"]):.1%}'
                    row += 1
            row += 1
        
    def _create_department_sheet(self, workbook: Workbook, stats: Dict):
        ws = workbook.create_sheet('部门结构')

        # 设置标题
        ws['A1'] = f'{self.current_month[:4]}年{self.current_month[4:]}月部门结构'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        # 部门人员统计
        department_stats = stats.get('部门结构', {})
        if not department_stats:
            ws['A3'] = '无部门数据'
            return
        row = 3
        ws[f'A{row}'] = '各部门人员统计'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = '部门名称'
        ws[f'B{row}'] = '人数'
        ws[f'C{row}'] = '占比'
        row += 1

        for dept, count in department_stats.items():
            ws[f'A{row}'] = dept
            ws[f'B{row}'] = count
            if stats['总人数'] > 0:
                ws[f'C{row}'] = f'{(count/stats["总人数"]):1.%}'
            row += 1